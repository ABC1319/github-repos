#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitHub Repos Exporter v2.0
导出 GitHub 个人仓库为 JSON / CSV / XLSX / HTML 单页导航
Fork & Modified from github-star: https://github.com/ABC1319/github-star
"""

import requests
import json
import csv
import os
import sys
import time
import argparse
import html as html_module
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

API_BASE = "https://api.github.com"


def get_headers(token):
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    return headers


def fetch_all_repos(username, token):
    """获取用户的所有公开仓库（使用 Search API 绕过未认证速率限制）"""
    repos = []
    page = 1
    per_page = 100
    headers = get_headers(token)
    print(f"[1/4] 正在获取用户 {username} 的仓库...")
    while True:
        url = f"{API_BASE}/search/repositories?q=user:{username}&per_page={per_page}&page={page}&sort=updated&order=desc"
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 404:
            print(f" 错误：用户 '{username}' 不存在")
            return []
        if resp.status_code == 403:
            print(f" 错误：API 速率限制。建议添加 GITHUB_TOKEN")
            return []
        resp.raise_for_status()
        data = resp.json()
        items = data.get("items", [])
        if not items:
            break
        repos.extend(items)
        print(f" 第 {page} 页: {len(items)} 个，累计 {len(repos)} 个")
        if len(items) < per_page:
            break
        page += 1
        time.sleep(0.2)
    print(f" 共获取 {len(repos)} 个仓库\n")
    return repos


def build_repo_data(repos):
    results = []
    total = len(repos)
    print(f"[2/4] 正在处理 {total} 个仓库...")
    for idx, repo in enumerate(repos, 1):
        item = {
            "序号": idx,
            "项目名称": repo.get("name", ""),
            "项目全名": repo.get("full_name", ""),
            "项目链接": repo.get("html_url", ""),
            "仓库描述": repo.get("description") or "",
            "主页网址": repo.get("homepage") or "",
            "项目语言": repo.get("language") or "",
            "星标数": repo.get("stargazers_count", 0),
            "Fork数": repo.get("forks_count", 0),
            "Watch数": repo.get("watchers_count", 0),
            "Open_Issues": repo.get("open_issues_count", 0),
            "默认分支": repo.get("default_branch", ""),
            "创建时间": repo.get("created_at", ""),
            "更新时间": repo.get("updated_at", ""),
            "推送时间": repo.get("pushed_at", ""),
            "仓库大小_KB": repo.get("size", 0),
            "是否为Fork": repo.get("fork", False),
            "Topics": ", ".join(repo.get("topics", [])),
            "License": (repo.get("license") or {}).get("spdx_id", ""),
            "Owner": repo.get("owner", {}).get("login", ""),
            "Owner类型": repo.get("owner", {}).get("type", ""),
            "Owner头像": repo.get("owner", {}).get("avatar_url", ""),
        }
        results.append(item)
        if idx % 50 == 0 or idx == total:
            print(f" [{idx}/{total}] 已处理")
    print(f" 全部处理完成\n")
    return results


def export_json(data, filepath):
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f" JSON: {filepath}")


def export_csv(data, filepath):
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    fieldnames = list(data[0].keys())
    csv_fieldnames = [f for f in fieldnames if f not in ["Owner头像"]]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow({k: v for k, v in row.items() if k in csv_fieldnames})
    print(f" CSV: {filepath}")


def export_xlsx(data, filepath):
    if not HAS_OPENPYXL:
        print(" 未安装 openpyxl，跳过 XLSX")
        return
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GitHub Repos"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="24292F", end_color="24292F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )
    fieldnames = list(data[0].keys())
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, row in enumerate(data, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = cell_align
            cell.border = thin_border
    col_widths = {
        "序号": 6, "项目名称": 25, "项目全名": 35, "项目链接": 45,
        "仓库描述": 50, "主页网址": 40, "项目语言": 12,
        "星标数": 10, "Fork数": 10, "Watch数": 10, "Open_Issues": 12,
        "默认分支": 12, "创建时间": 20, "更新时间": 20, "推送时间": 20,
        "仓库大小_KB": 14, "是否为Fork": 12, "Topics": 40, "License": 20,
        "Owner": 18, "Owner类型": 12, "Owner头像": 45
    }
    for col_idx, field in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f" XLSX: {filepath}")


def export_html(data, filepath, username):
    if not data:
        return
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)

    total_stars = sum(r.get("星标数", 0) for r in data)
    total_forks = sum(r.get("Fork数", 0) for r in data)
    lang_counts = {}
    for r in data:
        lang = r.get("项目语言") or "Unknown"
        lang_counts[lang] = lang_counts.get(lang, 0) + 1

    js_data = []
    for r in data:
        js_data.append({
            "name": r["项目名称"],
            "full_name": r["项目全名"],
            "html_url": r["项目链接"],
            "description": r["仓库描述"],
            "homepage": r["主页网址"],
            "language": r["项目语言"],
            "topics": r["Topics"],
            "stargazers_count": r["星标数"],
            "forks_count": r["Fork数"],
            "watchers_count": r["Watch数"],
            "open_issues_count": r["Open_Issues"],
            "license": r["License"],
            "created_at": r["创建时间"],
            "updated_at": r["更新时间"],
            "pushed_at": r["推送时间"],
            "size": r["仓库大小_KB"],
            "is_fork": r["是否为Fork"],
            "owner_login": r["Owner"],
            "owner_avatar": r["Owner头像"],
            "default_branch": r["默认分支"],
        })
    js_json = json.dumps(js_data, ensure_ascii=False)
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 使用 str.replace 避免花括号冲突
    html_template = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__USERNAME__ 的 GitHub 仓库导航</title>
<style>
:root {
  --bg: #f6f8fa;
  --card-bg: #ffffff;
  --text: #24292f;
  --text-secondary: #57606a;
  --border: #d0d7de;
  --accent: #0969da;
  --accent-hover: #0550ae;
  --star: #e3b341;
  --fork: #7d8590;
  --shadow: 0 1px 3px rgba(27,31,35,0.04), 0 1px 2px rgba(27,31,35,0.04);
  --radius: 12px;
}
[data-theme="dark"] {
  --bg: #0d1117;
  --card-bg: #161b22;
  --text: #c9d1d9;
  --text-secondary: #8b949e;
  --border: #30363d;
  --accent: #58a6ff;
  --accent-hover: #79c0ff;
  --star: #e3b341;
  --fork: #8b949e;
  --shadow: 0 1px 3px rgba(0,0,0,0.3);
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Noto Sans SC", Helvetica, Arial, sans-serif;
  background: var(--bg);
  color: var(--text);
  line-height: 1.6;
  transition: background 0.3s, color 0.3s;
}
.container { max-width: 1200px; margin: 0 auto; padding: 24px 16px; }
header {
  text-align: center;
  padding: 48px 16px 32px;
  position: relative;
}
header h1 { font-size: 2.2rem; font-weight: 700; margin-bottom: 8px; }
header p { color: var(--text-secondary); font-size: 1.05rem; }
.avatar {
  width: 80px; height: 80px; border-radius: 50%;
  border: 3px solid var(--border);
  margin-bottom: 16px;
  object-fit: cover;
}
.stats-bar {
  display: flex; justify-content: center; gap: 32px;
  margin: 24px 0; flex-wrap: wrap;
}
.stat-item { text-align: center; }
.stat-item .num { font-size: 1.6rem; font-weight: 700; color: var(--accent); }
.stat-item .label { font-size: 0.85rem; color: var(--text-secondary); }

.toolbar {
  display: flex; gap: 12px; flex-wrap: wrap;
  align-items: center; justify-content: center;
  margin: 24px 0 16px;
  padding: 16px;
  background: var(--card-bg);
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  position: sticky; top: 0; z-index: 100;
}
.search-box {
  position: relative; flex: 1; min-width: 240px; max-width: 400px;
}
.search-box input {
  width: 100%; padding: 10px 14px 10px 38px;
  border: 1px solid var(--border); border-radius: 8px;
  background: var(--bg); color: var(--text);
  font-size: 0.95rem; outline: none;
  transition: border-color 0.2s;
}
.search-box input:focus { border-color: var(--accent); }
.search-box svg {
  position: absolute; left: 12px; top: 50%; transform: translateY(-50%);
  width: 16px; height: 16px; fill: var(--text-secondary);
}
.search-hint {
  position: absolute; right: 10px; top: 50%; transform: translateY(-50%);
  font-size: 0.75rem; color: var(--text-secondary);
  background: var(--card-bg); padding: 2px 6px; border-radius: 4px;
  border: 1px solid var(--border);
}
select, button {
  padding: 10px 14px; border: 1px solid var(--border);
  border-radius: 8px; background: var(--bg); color: var(--text);
  font-size: 0.9rem; cursor: pointer; outline: none;
  transition: all 0.2s;
}
select:hover, button:hover { border-color: var(--accent); }
button.active { background: var(--accent); color: #fff; border-color: var(--accent); }

.lang-filters {
  display: flex; gap: 8px; flex-wrap: wrap; justify-content: center;
  margin-bottom: 24px;
}
.lang-chip {
  padding: 6px 14px; border-radius: 20px;
  border: 1px solid var(--border); background: var(--card-bg);
  font-size: 0.85rem; cursor: pointer; transition: all 0.2s;
  color: var(--text);
}
.lang-chip:hover { border-color: var(--accent); }
.lang-chip.active { background: var(--accent); color: #fff; border-color: var(--accent); }
.lang-chip .count { font-size: 0.75rem; opacity: 0.7; margin-left: 4px; }

.repo-grid {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
  gap: 16px;
}
.repo-card {
  background: var(--card-bg);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 20px;
  box-shadow: var(--shadow);
  transition: transform 0.2s, box-shadow 0.2s;
  display: flex; flex-direction: column;
}
.repo-card:hover {
  transform: translateY(-2px);
  box-shadow: 0 8px 24px rgba(0,0,0,0.08);
}
.repo-header {
  display: flex; align-items: flex-start; justify-content: space-between;
  margin-bottom: 10px;
}
.repo-name {
  font-size: 1.1rem; font-weight: 600;
  color: var(--accent); text-decoration: none;
  word-break: break-all;
}
.repo-name:hover { text-decoration: underline; }
.repo-fork {
  font-size: 0.75rem; color: var(--text-secondary);
  border: 1px solid var(--border); padding: 2px 8px;
  border-radius: 12px; margin-left: 8px; white-space: nowrap;
}
.repo-desc {
  color: var(--text-secondary); font-size: 0.9rem;
  margin-bottom: 12px; flex: 1;
  display: -webkit-box; -webkit-line-clamp: 3;
  -webkit-box-orient: vertical; overflow: hidden;
}
.repo-meta {
  display: flex; align-items: center; gap: 16px;
  flex-wrap: wrap; font-size: 0.85rem;
  color: var(--text-secondary);
}
.repo-meta span { display: inline-flex; align-items: center; gap: 4px; }
.lang-dot {
  width: 10px; height: 10px; border-radius: 50%;
  display: inline-block;
}
.repo-topics {
  display: flex; gap: 6px; flex-wrap: wrap;
  margin-top: 10px;
}
.topic-tag {
  font-size: 0.75rem; color: var(--accent);
  background: rgba(9,105,218,0.08);
  padding: 3px 10px; border-radius: 12px;
}
[data-theme="dark"] .topic-tag {
  background: rgba(88,166,255,0.12);
}
.highlight { background: rgba(9,105,218,0.15); border-radius: 2px; padding: 0 2px; }
[data-theme="dark"] .highlight { background: rgba(88,166,255,0.25); }

.empty-state {
  text-align: center; padding: 60px 20px;
  color: var(--text-secondary);
}
.empty-state svg { width: 64px; height: 64px; margin-bottom: 16px; fill: var(--border); }

#backTop {
  position: fixed; bottom: 24px; right: 24px;
  width: 44px; height: 44px; border-radius: 50%;
  background: var(--accent); color: #fff;
  border: none; cursor: pointer; display: none;
  align-items: center; justify-content: center;
  box-shadow: 0 4px 12px rgba(9,105,218,0.3);
  transition: opacity 0.3s, transform 0.2s;
  z-index: 200;
}
#backTop:hover { transform: scale(1.1); }
#backTop.show { display: flex; }

.theme-toggle {
  position: fixed; top: 16px; right: 16px;
  width: 40px; height: 40px; border-radius: 50%;
  background: var(--card-bg); border: 1px solid var(--border);
  cursor: pointer; display: flex; align-items: center; justify-content: center;
  box-shadow: var(--shadow); z-index: 200;
  transition: transform 0.2s;
}
.theme-toggle:hover { transform: rotate(20deg); }
.theme-toggle svg { width: 20px; height: 20px; fill: var(--text); }

.footer {
  text-align: center; padding: 40px 16px;
  color: var(--text-secondary); font-size: 0.85rem;
}
.footer a { color: var(--accent); text-decoration: none; }

@media (max-width: 640px) {
  header h1 { font-size: 1.6rem; }
  .toolbar { flex-direction: column; align-items: stretch; }
  .search-box { max-width: 100%; }
  .repo-grid { grid-template-columns: 1fr; }
  .stats-bar { gap: 20px; }
}
</style>
</head>
<body>
<button class="theme-toggle" onclick="toggleTheme()" title="切换主题">
  <svg id="moon-icon" viewBox="0 0 24 24"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"/></svg>
  <svg id="sun-icon" viewBox="0 0 24 24" style="display:none"><circle cx="12" cy="12" r="5"/><line x1="12" y1="1" x2="12" y2="3"/><line x1="12" y1="21" x2="12" y2="23"/><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"/><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"/><line x1="1" y1="12" x2="3" y2="12"/><line x1="21" y1="12" x2="23" y2="12"/><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"/><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"/></svg>
</button>

<div class="container">
<header>
  <img class="avatar" src="https://github.com/__USERNAME__.png" alt="__USERNAME__" onerror="this.style.display='none'">
  <h1>__USERNAME__ 的 GitHub 仓库</h1>
  <p>共 __TOTAL__ 个公开仓库 · 实时搜索 · 多维度排序 · 一键筛选</p>
  <div class="stats-bar">
    <div class="stat-item"><div class="num">__TOTAL__</div><div class="label">仓库</div></div>
    <div class="stat-item"><div class="num">__STARS__</div><div class="label">Stars</div></div>
    <div class="stat-item"><div class="num">__FORKS__</div><div class="label">Forks</div></div>
    <div class="stat-item"><div class="num">__LANGS__</div><div class="label">语言</div></div>
  </div>
</header>

<div class="toolbar">
  <div class="search-box">
    <svg viewBox="0 0 16 16"><path d="M10.68 11.74a6 6 0 0 1-7.922-8.982 6 6 0 0 1 8.982 7.922l3.04 3.04a.749.749 0 0 1-.326 1.275.749.749 0 0 1-.734-.215ZM11 6.5a4.499 4.499 0 1 0-8.997 0A4.499 4.499 0 0 0 11 6.5Z"/></svg>
    <input type="text" id="searchInput" placeholder="搜索仓库名称、描述、Topics、语言..." autocomplete="off">
    <span class="search-hint">Ctrl+K</span>
  </div>
  <select id="sortSelect">
    <option value="updated">默认（更新时间）</option>
    <option value="stars">星标数</option>
    <option value="name">名称</option>
    <option value="created">创建时间</option>
  </select>
</div>

<div class="lang-filters" id="langFilters"></div>

<div id="repoGrid" class="repo-grid"></div>

<div class="empty-state" id="emptyState" style="display:none">
  <svg viewBox="0 0 24 24"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-2 15l-5-5 1.41-1.41L10 14.17l7.59-7.59L19 8l-9 9z"/></svg>
  <p>没有找到匹配的仓库</p>
</div>

<div class="footer">
  <p>由 <a href="https://github.com/__USERNAME__" target="_blank">@__USERNAME__</a> 生成 · 
  <a href="https://github.com/__USERNAME__/github-repos" target="_blank">github-repos</a> 风格</p>
  <p style="margin-top:4px;opacity:0.7">生成时间：__TIME__</p>
</div>
</div>

<button id="backTop" onclick="window.scrollTo({top:0,behavior:'smooth'})">
  <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M18 15l-6-6-6 6"/></svg>
</button>

<script>
const repos = __JS_DATA__;
const langColors = {
  JavaScript: '#f1e05a', TypeScript: '#2b7489', Python: '#3572A5', HTML: '#e34c26',
  CSS: '#563d7c', Java: '#b07219', 'C++': '#f34b7d', C: '#555555', 'C#': '#178600',
  Go: '#00ADD8', Rust: '#dea584', Ruby: '#701516', PHP: '#4F5D95', Swift: '#ffac45',
  Kotlin: '#A97BFF', Vue: '#41b883', Shell: '#89e051', Dart: '#00B4AB',
  R: '#198CE7', Scala: '#c22d40', Perl: '#0298c3', Lua: '#000080',
  Jupyter: '#DA5B0B', Dockerfile: '#384d54'
};
let currentLang = 'all';
let currentSort = 'updated';
let currentQuery = '';

function escapeHtml(s) {
  return s.replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}
function highlight(text, query) {
  if (!query) return escapeHtml(text);
  const q = query.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
  return escapeHtml(text).replace(new RegExp(`(${q})`, 'gi'), '<span class="highlight">$1</span>');
}
function formatDate(d) {
  if (!d) return '';
  return d.slice(0, 10);
}
function getLangColor(lang) {
  return langColors[lang] || '#8b949e';
}

function renderLangFilters() {
  const counts = {};
  repos.forEach(r => { const l = r.language || 'Unknown'; counts[l] = (counts[l]||0)+1; });
  const sorted = Object.entries(counts).sort((a,b) => b[1]-a[1]);
  const container = document.getElementById('langFilters');
  let html = `<button class="lang-chip ${currentLang==='all'?'active':''}" onclick="setLang('all')">全部<span class="count">${repos.length}</span></button>`;
  sorted.forEach(([lang, count]) => {
    html += `<button class="lang-chip ${currentLang===lang?'active':''}" onclick="setLang('${escapeHtml(lang)}')">${escapeHtml(lang)}<span class="count">${count}</span></button>`;
  });
  container.innerHTML = html;
}

function getFiltered() {
  let list = repos.filter(r => {
    if (currentLang !== 'all' && (r.language||'Unknown') !== currentLang) return false;
    if (!currentQuery) return true;
    const q = currentQuery.toLowerCase();
    return (r.name||'').toLowerCase().includes(q) ||
           (r.description||'').toLowerCase().includes(q) ||
           (r.topics||'').toLowerCase().includes(q) ||
           (r.language||'').toLowerCase().includes(q) ||
           (r.owner_login||'').toLowerCase().includes(q);
  });
  list.sort((a,b) => {
    if (currentSort === 'stars') return b.stargazers_count - a.stargazers_count;
    if (currentSort === 'name') return a.name.localeCompare(b.name);
    if (currentSort === 'created') return b.created_at.localeCompare(a.created_at);
    return b.updated_at.localeCompare(a.updated_at);
  });
  return list;
}

function render() {
  const list = getFiltered();
  const grid = document.getElementById('repoGrid');
  const empty = document.getElementById('emptyState');
  if (list.length === 0) {
    grid.innerHTML = '';
    empty.style.display = 'block';
    return;
  }
  empty.style.display = 'none';
  grid.innerHTML = list.map(r => {
    const topics = r.topics ? r.topics.split(', ').filter(t=>t) : [];
    const forkBadge = r.is_fork ? `<span class="repo-fork">Fork</span>` : '';
    const langDot = r.language ? `<span class="lang-dot" style="background:${getLangColor(r.language)}"></span>` : '';
    return `<div class="repo-card">
      <div class="repo-header">
        <div style="display:flex;align-items:center;flex-wrap:wrap;gap:4px;">
          <a class="repo-name" href="${r.html_url}" target="_blank" rel="noopener">${highlight(r.name, currentQuery)}</a>
          ${forkBadge}
        </div>
      </div>
      <div class="repo-desc">${r.description ? highlight(r.description, currentQuery) : '<i style="opacity:0.5">暂无描述</i>'}</div>
      <div class="repo-meta">
        ${r.language ? `<span>${langDot}${highlight(r.language, currentQuery)}</span>` : ''}
        <span>⭐ ${r.stargazers_count}</span>
        <span>🍴 ${r.forks_count}</span>
        <span>📅 ${formatDate(r.updated_at)}</span>
      </div>
      ${topics.length ? `<div class="repo-topics">${topics.map(t => `<span class="topic-tag">${highlight(t, currentQuery)}</span>`).join('')}</div>` : ''}
    </div>`;
  }).join('');
}

function setLang(lang) {
  currentLang = lang;
  renderLangFilters();
  render();
}

document.getElementById('searchInput').addEventListener('input', e => {
  currentQuery = e.target.value;
  render();
});
document.getElementById('sortSelect').addEventListener('change', e => {
  currentSort = e.target.value;
  render();
});
document.addEventListener('keydown', e => {
  if ((e.ctrlKey || e.metaKey) && e.key === 'k') {
    e.preventDefault();
    document.getElementById('searchInput').focus();
  }
  if (e.key === 'Escape') {
    document.getElementById('searchInput').value = '';
    currentQuery = '';
    render();
  }
});

window.addEventListener('scroll', () => {
  const btn = document.getElementById('backTop');
  if (window.scrollY > 300) btn.classList.add('show');
  else btn.classList.remove('show');
});

function toggleTheme() {
  const html = document.documentElement;
  const isDark = html.getAttribute('data-theme') === 'dark';
  html.setAttribute('data-theme', isDark ? 'light' : 'dark');
  localStorage.setItem('theme', isDark ? 'light' : 'dark');
  document.getElementById('moon-icon').style.display = isDark ? 'block' : 'none';
  document.getElementById('sun-icon').style.display = isDark ? 'none' : 'block';
}
(function initTheme() {
  const saved = localStorage.getItem('theme');
  const prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
  const isDark = saved ? saved === 'dark' : prefersDark;
  if (isDark) {
    document.documentElement.setAttribute('data-theme', 'dark');
    document.getElementById('moon-icon').style.display = 'none';
    document.getElementById('sun-icon').style.display = 'block';
  }
})();

renderLangFilters();
render();
</script>
</body>
</html>"""

    html_content = (html_template
        .replace("__USERNAME__", username)
        .replace("__TOTAL__", str(len(data)))
        .replace("__STARS__", str(total_stars))
        .replace("__FORKS__", str(total_forks))
        .replace("__LANGS__", str(len(lang_counts)))
        .replace("__TIME__", gen_time)
        .replace("__JS_DATA__", js_json))

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f" HTML: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="导出 GitHub 个人仓库")
    parser.add_argument("--output", "-o", default="./dist", help="输出目录 (默认: ./dist)")
    parser.add_argument("--username", "-u", default=os.environ.get("GITHUB_USERNAME"), help="GitHub 用户名")
    parser.add_argument("--token", "-t", default=os.environ.get("GITHUB_TOKEN"), help="GitHub Token")
    parser.add_argument("--formats", "-f", default="json,csv,xlsx,html", help="导出格式，逗号分隔")
    args = parser.parse_args()

    username = args.username
    token = args.token
    output_dir = args.output
    formats = [f.strip().lower() for f in args.formats.split(",")]

    if not username:
        print("错误：未指定 GitHub 用户名。请通过 --username 参数或 GITHUB_USERNAME 环境变量设置。")
        sys.exit(1)

    print(f"输出目录: {output_dir}")
    print(f"导出格式: {formats}\n")

    repos = fetch_all_repos(username, token)
    if not repos:
        print("未获取到任何仓库，退出。")
        sys.exit(1)

    data = build_repo_data(repos)

    print("[3/4] 正在导出文件...")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"github_repos_{username}_{timestamp}"

    for fmt in formats:
        if fmt == "json":
            export_json(data, os.path.join(output_dir, f"{base_name}.json"))
            export_json(data, os.path.join(output_dir, "repos.json"))
        elif fmt == "csv":
            export_csv(data, os.path.join(output_dir, f"{base_name}.csv"))
            export_csv(data, os.path.join(output_dir, "repos.csv"))
        elif fmt == "xlsx":
            export_xlsx(data, os.path.join(output_dir, f"{base_name}.xlsx"))
            export_xlsx(data, os.path.join(output_dir, "repos.xlsx"))
        elif fmt == "html":
            export_html(data, os.path.join(output_dir, "index.html"), username)
        else:
            print(f" 未知格式: {fmt}")

    print("\n[4/4] 全部完成！")
    print(f" 输出目录: {output_dir}")
    for f in sorted(os.listdir(output_dir)):
        print(f" - {f}")


if __name__ == "__main__":
    main()
